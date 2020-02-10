from selenium import webdriver
import time
from pprint import pprint
import openpyxl

d = {}
url_list = []
wd = webdriver.Chrome()

isDebug = False
debugUrl = 'https://gzmingtai.en.alibaba.com/company_profile.html?spm=a2700.galleryofferlist.0.0.7d4b6a3a4AYF71#top-nav-bar'

total_num = 1

# 起始位置
start_category = 0
start_page = 1
# 顺序：True
# 逆序: False
order = True

category = -1
cur_page = -1

output_num = 1
curRow = 1
maxRow = 5000
sheetname = 'Sheet'
result = openpyxl.Workbook()
result.create_sheet(sheetname)
result_sheet = result[sheetname]


section1 = [
    'Business Type',
    'Country / Region',
    'Main Products',
    'Total Employees',
    'Total Annual Revenue',
    'Year Established',
    'Certifications',
    'Product Certifications',
    'Patents',
    'Trademarks',
    'Main Markets'
]

section2 = [
    'Registration No.:',
    'Company Name:',
    'Date of Issue:',
    'Date of Expiry:',
    'Registered Capital:',

    'Operational Address:',
    'Total Export Revenue:',
    'Export Percentage:'
]

single = [
    'firm name',
    'GS year',
    'transaction quantity',
    'transaction value',
    'Business Type',
    'Main Products',
    'Total Employees',
    'Year Established',
    'Total Annual Revenue',
    'Main Markets',
    'Country / Region',
    'Ownership', 
    'Factory Size',
    'Factory Country/Region',
    'No. of Production Lines',
    'Contract Manufacturing',
    'Annual Output Value',
    'Main Product(s)', # is Main Products ?
    'Total Annual Revenue:',
    'Date of Issue:',
    'Date of Expiry:',
    'Operational Address:',
    'Registered Capital:',
    'Total Export Revenue:',
    'Export Percentage:'
]
single2list = [
    'Certifications',
    'Product Certifications',
    'Patents',
    'Trademarks'
]
unuse = [
    'Test Equipment',
    'Machine Name',
    'Brand & Model NO',
    'Quantity',
    'Awards Certification	Name',
    'Issued By',
    'Start Date',
    'Description',
    'Main Markets',
    'Total Revenue(%)',
    'Main Product(s)'
]

mul2mul_judge = [
    'Production Equipment',
    'Annual Production Capacity',
    'Production Certification',
    'Certification',
    'Trademarks',
    'Patents'
]

mul2mul = {
    'Production Equipment': [
        'Name',
        'No',
        'Quantity'
    ],

    'Annual Production Capacity': [
        'Product Name',
        'Production Line Capacity',
        'Actual Units Produced(Previous Year)'
    ],
    'Production Certification': [
        'Certification Name',
        'Certified By',
        'Business Scope',
        'Available Date'
    ],
    'Certification': [
        'Certification Name',
        'Certified By',
        'Business Scope',
        'Available Date'
    ],
    'Trademarksmul2mul': [
        'Trademark No',
        'Trademark Name',
        'Trademark Category',
        'Available Date'
    ],
    'Patentsmul2mul': [
        'Patent No',
        'Patent Name',
        'Patent Type',
        'Available Date'
    ]
}


def get_url_list():
    global url_list
    wd.get('https://www.alibaba.com/Products')
    s = []
    for element in wd.find_elements_by_class_name('sub-item-cont'):
        for temp in element.find_elements_by_tag_name('a'):
            url_list.append(temp.get_attribute('href'))


def spider(url):
    global start_page
    global cur_page
    cur_start_page = start_page if start_page > 0 else 1
    for cur_page in range(cur_start_page, 101):
        realUrl = url + '?page={0}'.format(cur_page)
        wd.get(realUrl)
        wd.execute_script('window.scrollTo(0, document.body.scrollHeight)')
        time.sleep(1)
        company_urls = []
        if "did not match any products" in wd.page_source:
            print('本类已结束')
            break
        for company in wd.find_elements_by_class_name('organic-gallery-offer__seller-company'):
            company_urls.append(company.get_attribute('href'))
        for company_url in company_urls:
            get_company_info(company_url)


def get_company_info(url):
    global d
    global total_num
    global category
    global cur_page
    print('正在处理第{0}类，第{1}页, 总第{2}个数据'.format(category, cur_page, total_num))
    initD()
    error_time = 0
    while True:
        try:
            wd.get(url)
            if not ('join-year' in wd.page_source):
                print('第{0}类，第{1}页, 总第{2}个数据被忽略，对应链接{3}'.format(category, cur_page, total_num, url))
                total_num += 1
                return
            d['GS year'] = wd.find_element_by_class_name('join-year').find_element_by_class_name('value').get_attribute('textContent').strip() + 'YEARS'
            d['firm name'] = wd.find_element_by_class_name('company-info').find_element_by_class_name('title-text').get_attribute('textContent').strip()
            quantity = wd.find_elements_by_class_name('transaction-detail-title')[1].get_attribute('textContent').strip().split(' ')
            d['transaction quantity'] = '' if len(quantity) == 1 else quantity[0]
            d['transaction value'] = wd.find_elements_by_class_name('transaction-detail-content')[1].get_attribute('textContent').strip()

            basicInfo = wd.find_element_by_class_name('company-basicInfo')
            for key, value in zip(
                    basicInfo.find_elements_by_class_name('field-title'),
                    basicInfo.find_elements_by_class_name('content-value')
            ):
                text = key.get_attribute('textContent').strip()
                if '(' in text:
                    text = text.split('(')[0]
                d[text] = value.get_attribute('textContent').strip()
            if 'infoList-mod-field' in wd.page_source:
                fields = wd.find_elements_by_class_name('infoList-mod-field')
                for field in fields:
                    title = field.find_element_by_tag_name('h3').get_attribute('textContent').strip()
                    if title == 'Factory Information':
                        for item in field.find_elements_by_class_name('icbu-shop-table-col-item'):
                            key_value = item.find_elements_by_tag_name('span')
                            key = key_value[0].get_attribute('textContent').strip()
                            value = key_value[1].get_attribute('textContent').strip()
                            d[key] = value
                    elif title in mul2mul_judge:
                        if title + 'mul2mul' in mul2mul:
                            title = title + 'mul2mul'
                        lists = mul2mul[title]
                        keys = field.find_element_by_class_name('next-table-header').find_elements_by_class_name('next-table-cell-wrapper')
                        eachs = field.find_element_by_class_name('next-table-body').find_elements_by_class_name('next-table-row')
                        for each in eachs:
                            d_temp = {}
                            for li in lists:
                                d_temp[li] = ''
                            values = each.find_elements_by_class_name('next-table-cell-wrapper')
                            for key, value in zip(keys, values):
                                if key.get_attribute('textContent').strip() in lists:
                                    d_temp[key.get_attribute('textContent').strip()] = value.get_attribute('textContent').strip()
                            d[title].append(d_temp)
            profile_url = url.split(".html")[0] + '/trustpass_profile.html'
            trade_url = url.split(".html")[0] + '/trade_capacity.html'
            wd.get(profile_url)
            table = wd.find_elements_by_class_name('table')
            if 'Registration No' not in wd.page_source:
                print('第{0}类，第{1}页, 总第{2}个数据被忽略，对应链接{3}'.format(category, cur_page, total_num, url))
                total_num += 1
                return
            if len(table) > 0:
                tab = table[0]
                flag = False
                for item in tab.find_elements_by_tag_name('tr'):
                    key = item.find_element_by_tag_name('th').get_attribute('textContent').strip()
                    value = item.find_element_by_tag_name('td').get_attribute('textContent').strip()
                    d[key] = value
                    if key == 'Registration No.:':
                        flag = True
                if not flag:
                    keys_ele = tab.find_elements_by_tag_name('dt')
                    values_ele = tab.find_elements_by_tag_name('dd')
                    for key_ele, value_ele in zip(keys_ele, values_ele):
                        if key_ele.get_attribute('textContent').strip() in section2:
                            d[key_ele.get_attribute('textContent').strip()] = value_ele.get_attribute('textContent').strip()
                    
            wd.get(trade_url)
            if 'article' in wd.page_source:
                table = wd.find_element_by_class_name('article').find_element_by_class_name('table')
                keys_ele = table.find_elements_by_tag_name('th')
                values_ele = table.find_elements_by_tag_name('td')
                for key, value in zip(keys_ele, values_ele):
                    if key.get_attribute('textContent').strip() in section2:
                        d[key.get_attribute('textContent').strip()] = value.get_attribute('textContent').strip()
            break
        except Exception as e:
            if error_time >= 3:
                print('第{0}类，第{1}页, 总第{2}个数据被忽略，对应链接{3}'.format(category, cur_page, total_num, url))
                if isDebug:
                    pprint(d)
                    exit(0)
                total_num += 1
                return
            else:
                error_time += 1
    if isDebug:
        pprint(d)
        exit(0)
    total_num += 1
    wirteD(d)



def initD():
    global d
    for item in single:
        d[item] = ''
    for item in single2list:
        d[item] = ''
    for item in mul2mul.keys():
        d[item] = []
    d['Registration No.:'] = ''
    d['Company Name:'] = ''


def wirteD(d):
    global curRow
    global result
    global result_sheet
    global output_num
    global category
    if curRow == maxRow:
        result.save('./' + '{0}.xlsx'.format(output_num))
        print('已将该部分数据写入文件{0}.xlsx'.format(output_num))
        result.close()
        output_num += 1
        curRow = 1
        result = openpyxl.Workbook()
        result.create_sheet(sheetname)
        result_sheet = result[sheetname]
    curRow += 1
    j = 1
    try:
        for item in single:
            result_sheet.cell(row=curRow, column=j).value = d[item]
            j += 1
        for item in single2list:
            result_sheet.cell(row=curRow, column=j).value = d[item]
            j += 1
        result_sheet.cell(row=curRow, column=j).value = d['Registration No.:']
        result_sheet.cell(row=curRow, column=j + 1).value = d['Company Name:']
        j += 2
        for item in mul2mul.keys():
            for d_temp in d[item]:
                for s in mul2mul[item]:
                    result_sheet.cell(
                        row=curRow, column=j).value = d_temp[s]
                    j = j + 1
            j += 3*(15 - len(d[item]))
    except Exception as e:
        print('保存数据时出现错误，输出文件中')
        print(curRow)
        print(j)
        result_sheet.cell(row=curRow, column=j).value = 'error'
        pprint(d)
        result.save('./{0}.xlsx'.format(output_num))
        print(e)
        exit(0)


if __name__ == '__main__':
    wd.set_window_size(1000, 30000)
    wd.implicitly_wait(4)
    if isDebug:
        get_company_info(debugUrl)
        exit(0)
    get_url_list()
    
    ran = []
    if order:
        ran = range(start_category, len(url_list))
    else:
        ran = range(len(url_list) - 1, -1, -1)
    
    for i in ran:
        try:
            category = i
            spider(url_list[i])
        except:
            result.save('./{0}.xlsx'.format(output_num))
            print('已将该部分数据写入文件{0}.xlsx'.format(output_num))
            exit(0)